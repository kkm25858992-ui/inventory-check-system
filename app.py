from flask import (
    Flask,
    render_template,
    request,
    send_file,
    session,
    redirect,
    jsonify
)

import pandas as pd
import uuid
import os
import time
import qrcode

from io import BytesIO
from werkzeug.utils import secure_filename
from openpyxl import load_workbook


# =========================================================
# Flask 설정
# =========================================================

app = Flask(__name__)

app.secret_key = "secret_key_123"


# =========================================================
# 계정
# =========================================================

admins = {
    "김경민": "ourbox123"
}

users = {
    "김경민": "ourbox",
    "8층": "1234",
    "7층": "5678"
}


# =========================================================
# 파일 설정
# =========================================================

UPLOAD_FOLDER = "files"

os.makedirs(
    UPLOAD_FOLDER,
    exist_ok=True
)

FILE_EXPIRE_TIME = 60 * 60


# =========================================================
# 오래된 파일 삭제
# =========================================================

def delete_old_files():

    now = time.time()

    for filename in os.listdir(UPLOAD_FOLDER):

        file_path = os.path.join(
            UPLOAD_FOLDER,
            filename
        )

        if not os.path.isfile(file_path):
            continue

        try:

            file_age = (
                now -
                os.path.getmtime(file_path)
            )

            if file_age > FILE_EXPIRE_TIME:

                os.remove(file_path)

        except Exception:
            pass


# =========================================================
# 숫자 정리
# =========================================================

def clean_number(value):

    if value is None:
        return 0

    try:

        text = str(value)

        text = text.replace(",", "").strip()

        if text == "":
            return 0

        number = float(text)

        if number.is_integer():
            return int(number)

        return number

    except Exception:

        return 0


# =========================================================
# 문자열 정리
# =========================================================

def clean_string(value):

    if value is None:
        return ""

    try:

        if pd.isna(value):
            return ""

    except Exception:
        pass

    return str(value).strip()


# =========================================================
# 로그인 페이지
# =========================================================

@app.route("/login")
def login_page():

    return render_template(
        "login.html"
    )


# =========================================================
# 로그인
# =========================================================

@app.route(
    "/login",
    methods=["POST"]
)
def login():

    user_id = request.form.get(
        "id",
        ""
    )

    pw = request.form.get(
        "pw",
        ""
    )

    role = request.form.get(
        "role",
        ""
    )

    # 관리자
    if role == "admin":

        if (
            user_id in admins
            and admins[user_id] == pw
        ):

            session["login"] = True
            session["role"] = "admin"

            return redirect("/admin")

    # 사용자
    if role == "user":

        if (
            user_id in users
            and users[user_id] == pw
        ):

            session["login"] = True
            session["role"] = "user"

            return redirect("/")

    return "로그인 실패"


# =========================================================
# 사용자 페이지
# =========================================================

@app.route("/")
def index():

    if not session.get("login"):
        return redirect("/login")

    if session.get("role") != "user":
        return redirect("/login")

    return render_template(
        "index.html",
        data=[],
        mapping=[]
    )


# =========================================================
# 관리자 페이지
# =========================================================

@app.route("/admin")
def admin():

    if not session.get("login"):
        return redirect("/login")

    if session.get("role") != "admin":
        return redirect("/login")

    files = []

    for filename in os.listdir(UPLOAD_FOLDER):

        if not filename.endswith(".xlsx"):
            continue

        file_path = os.path.join(
            UPLOAD_FOLDER,
            filename
        )

        files.append({
            "id": filename.replace(
                ".xlsx",
                ""
            ),
            "time": time.strftime(
                "%Y-%m-%d %H:%M:%S",
                time.localtime(
                    os.path.getmtime(
                        file_path
                    )
                )
            )
        })

    files = sorted(
        files,
        key=lambda x: x["time"],
        reverse=True
    )

    return render_template(
        "admin.html",
        files=files
    )


# =========================================================
# 엑셀 업로드
# =========================================================

@app.route(
    "/upload",
    methods=["POST"]
)
def upload():

    try:

        file = request.files.get("file")

        if file is None:
            return "파일이 없습니다."

        if file.filename == "":
            return "파일을 선택해주세요."

        filename = secure_filename(
            file.filename.lower()
        )

        # =================================================
        # CSV
        # =================================================

        if filename.endswith(".csv"):

            df = pd.read_csv(file)

            mapping_df = df.copy()

        # =================================================
        # Excel
        # =================================================

        else:

            excel_data = pd.read_excel(
                file,
                sheet_name=None,
                engine="openpyxl"
            )

            if not excel_data:
                return "엑셀 데이터가 없습니다."

            sheet_names = list(
                excel_data.keys()
            )

            # 첫 번째 시트
            df = excel_data[
                sheet_names[0]
            ].copy()

            # 두 번째 시트
            if len(sheet_names) >= 2:

                mapping_df = excel_data[
                    sheet_names[1]
                ].copy()

            else:

                mapping_df = pd.DataFrame()

        # =================================================
        # 시트2 확인
        # =================================================

        if mapping_df.empty:

            return (
                "시트2에 상품 마스터 데이터가 없습니다."
            )

        # =================================================
        # 시트2 구조
        #
        # A 화주사
        # B 바코드
        # C 입수량
        # D 상품명
        # =================================================

        if all(
            col in mapping_df.columns
            for col in [
                "화주사",
                "바코드",
                "입수량",
                "상품명"
            ]
        ):

            mapping_df = mapping_df[
                [
                    "화주사",
                    "바코드",
                    "입수량",
                    "상품명"
                ]
            ].copy()

        elif len(mapping_df.columns) >= 4:

            mapping_df = mapping_df.iloc[
                :,
                0:4
            ].copy()

            mapping_df.columns = [
                "화주사",
                "바코드",
                "입수량",
                "상품명"
            ]

        else:

            return (
                "시트2는 "
                "A=화주사, "
                "B=바코드, "
                "C=입수량, "
                "D=상품명 "
                "구조여야 합니다."
            )

        # =================================================
        # 데이터 정리
        # =================================================

        mapping_df["화주사"] = (
            mapping_df["화주사"]
            .apply(clean_string)
        )

        mapping_df["바코드"] = (
            mapping_df["바코드"]
            .apply(clean_string)
        )

        mapping_df["입수량"] = (
            mapping_df["입수량"]
            .apply(clean_number)
        )

        mapping_df["상품명"] = (
            mapping_df["상품명"]
            .apply(clean_string)
        )

        # 빈 바코드 제거
        mapping_df = mapping_df[
            mapping_df["바코드"] != ""
        ].copy()

        # =================================================
        # JS로 전달
        # =================================================

        mapping = mapping_df.to_dict(
            orient="records"
        )

        return render_template(
            "index.html",
            data=[],
            mapping=mapping
        )

    except Exception as e:

        print(
            "UPLOAD ERROR:",
            e
        )

        return (
            "엑셀 업로드 오류: " +
            str(e)
        )


# =========================================================
# 저장
# =========================================================

@app.route(
    "/save",
    methods=["POST"]
)
def save():

    delete_old_files()

    try:

        body = request.get_json(
            silent=True
        )

        if not body:

            return jsonify({
                "error":
                    "전송된 데이터가 없습니다."
            }), 400

        # =================================================
        # 재고조사 데이터
        # =================================================

        inventory = body.get(
            "inventory",
            []
        )

        # =================================================
        # 상품 마스터
        # =================================================

        mapping = body.get(
            "mapping",
            []
        )

        if not isinstance(
            inventory,
            list
        ):

            inventory = []

        if not isinstance(
            mapping,
            list
        ):

            mapping = []

        if len(inventory) == 0:

            return jsonify({
                "error":
                    "재고조사 데이터가 없습니다."
            }), 400

        # =================================================
        # 시트1
        #
        # A 바코드
        # B 랙
        # C 소비기한
        # D 수량
        # E 상품명
        # F 화주사
        # =================================================

        sheet1_rows = []

        for item in inventory:

            if not isinstance(
                item,
                dict
            ):
                continue

            sheet1_rows.append({

                "바코드":
                    clean_string(
                        item.get(
                            "바코드",
                            ""
                        )
                    ),

                "랙":
                    clean_string(
                        item.get(
                            "랙",
                            ""
                        )
                    ),

                "소비기한":
                    clean_string(
                        item.get(
                            "소비기한",
                            ""
                        )
                    ),

                "수량":
                    clean_number(
                        item.get(
                            "수량",
                            0
                        )
                    ),

                "상품명":
                    clean_string(
                        item.get(
                            "상품명",
                            ""
                        )
                    ),

                "화주사":
                    clean_string(
                        item.get(
                            "화주사",
                            ""
                        )
                    )
            })

        df1 = pd.DataFrame(
            sheet1_rows,
            columns=[
                "바코드",
                "랙",
                "소비기한",
                "수량",
                "상품명",
                "화주사"
            ]
        )

        # =================================================
        # 시트2
        #
        # A 화주사
        # B 바코드
        # C 입수량
        # D 상품명
        # =================================================

        mapping_rows = []

        for item in mapping:

            if not isinstance(
                item,
                dict
            ):
                continue

            mapping_rows.append({

                "화주사":
                    clean_string(
                        item.get(
                            "화주사",
                            ""
                        )
                    ),

                "바코드":
                    clean_string(
                        item.get(
                            "바코드",
                            ""
                        )
                    ),

                "입수량":
                    clean_number(
                        item.get(
                            "입수량",
                            0
                        )
                    ),

                "상품명":
                    clean_string(
                        item.get(
                            "상품명",
                            ""
                        )
                    )
            })

        df2 = pd.DataFrame(
            mapping_rows,
            columns=[
                "화주사",
                "바코드",
                "입수량",
                "상품명"
            ]
        )

        # =================================================
        # 파일 ID
        # =================================================

        file_id = str(
            uuid.uuid4()
        )

        path = os.path.join(
            UPLOAD_FOLDER,
            f"{file_id}.xlsx"
        )

        # =================================================
        # 엑셀 생성
        # =================================================

        with pd.ExcelWriter(
            path,
            engine="openpyxl"
        ) as writer:

            # 시트1
            df1.to_excel(
                writer,
                index=False,
                sheet_name="시트1"
            )

            # 시트2
            df2.to_excel(
                writer,
                index=False,
                sheet_name="시트2"
            )

        # =================================================
        # 엑셀 서식
        # =================================================

        format_excel(path)

        return jsonify({
            "file_id": file_id
        })

    except Exception as e:

        print(
            "SAVE ERROR:",
            e
        )

        return jsonify({
            "error": str(e)
        }), 500


# =========================================================
# 엑셀 서식
# =========================================================

def format_excel(path):

    try:

        wb = load_workbook(path)

        # =================================================
        # 시트1
        # =================================================

        if "시트1" in wb.sheetnames:

            ws1 = wb["시트1"]

            widths1 = {
                "A": 22,
                "B": 18,
                "C": 16,
                "D": 12,
                "E": 30,
                "F": 20
            }

            for col, width in widths1.items():

                ws1.column_dimensions[
                    col
                ].width = width

            ws1.freeze_panes = "A2"

            if ws1.max_row >= 1:

                ws1.auto_filter.ref = (
                    ws1.dimensions
                )

        # =================================================
        # 시트2
        # =================================================

        if "시트2" in wb.sheetnames:

            ws2 = wb["시트2"]

            widths2 = {
                "A": 20,
                "B": 22,
                "C": 12,
                "D": 30
            }

            for col, width in widths2.items():

                ws2.column_dimensions[
                    col
                ].width = width

            ws2.freeze_panes = "A2"

            if ws2.max_row >= 1:

                ws2.auto_filter.ref = (
                    ws2.dimensions
                )

        wb.save(path)

    except Exception as e:

        print(
            "EXCEL FORMAT ERROR:",
            e
        )


# =========================================================
# 다운로드
# =========================================================

@app.route(
    "/download/<file_id>"
)
def download(file_id):

    file_id = secure_filename(
        file_id
    )

    path = os.path.join(
        UPLOAD_FOLDER,
        f"{file_id}.xlsx"
    )

    if not os.path.exists(path):

        return "파일 없음"

    return send_file(
        path,
        download_name="재고조사결과.xlsx",
        as_attachment=True
    )


# =========================================================
# 공유 다운로드
# =========================================================

@app.route(
    "/share/<file_id>"
)
def share_download(file_id):

    file_id = secure_filename(
        file_id
    )

    path = os.path.join(
        UPLOAD_FOLDER,
        f"{file_id}.xlsx"
    )

    if not os.path.exists(path):

        return "파일 없음"

    return send_file(
        path,
        download_name="재고조사결과.xlsx",
        as_attachment=True
    )


# =========================================================
# QR 생성
# =========================================================

@app.route(
    "/qr/<file_id>"
)
def generate_qr(file_id):

    file_id = secure_filename(
        file_id
    )

    path = os.path.join(
        UPLOAD_FOLDER,
        f"{file_id}.xlsx"
    )

    if not os.path.exists(path):

        return "파일 없음"

    url = (
        request.host_url.rstrip("/")
        + "/share/"
        + file_id
    )

    qr = qrcode.make(url)

    img_io = BytesIO()

    qr.save(
        img_io,
        "PNG"
    )

    img_io.seek(0)

    return send_file(
        img_io,
        mimetype="image/png"
    )


# =========================================================
# 파일 삭제
# =========================================================

@app.route(
    "/delete/<file_id>",
    methods=["POST"]
)
def delete_file(file_id):

    file_id = secure_filename(
        file_id
    )

    path = os.path.join(
        UPLOAD_FOLDER,
        f"{file_id}.xlsx"
    )

    if os.path.exists(path):

        try:

            os.remove(path)

            return "삭제 완료"

        except Exception as e:

            return (
                "삭제 실패: " +
                str(e)
            )

    return "파일 없음"


# =========================================================
# 로그아웃
# =========================================================

@app.route("/logout")
def logout():

    session.clear()

    return redirect("/login")


# =========================================================
# 실행
# =========================================================

if __name__ == "__main__":

    port = int(
        os.environ.get(
            "PORT",
            5000
        )
    )

    app.run(
        host="0.0.0.0",
        port=port,
        debug=False
    )
